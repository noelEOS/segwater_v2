"""Append a Hampyeong manuscript-estimator tab to the results registry (.xlsx).

Adds one tab holding the DEPRECATED pixel-bootstrap accuracy tables (6 metrics, 4 models,
3 dates, with 95% percentile CIs) computed by hampyeong_manuscript_estimator.py. The tab
carries a prominent banner: these numbers are for co-author REPORTING ONLY -- the method
is the superseded 400-pixel-subsample estimator, and the defensible inference lives in
STATISTICAL_ASSESSMENT.md.

Idempotent: if the tab already exists it is replaced. The rest of the workbook is untouched.

Run: /opt/homebrew/Caskroom/miniforge/base/envs/eda/bin/python <this file>
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REGISTRY = Path("experiments/demak_semarang/results_registry/demak_semarang_results_registry.xlsx")
CSV = Path("experiments/hampyeong/evaluation/manuscript_estimator_all_metrics.csv")
TAB = "hampyeong_estimator_REPORTING"

MODELS = ["legacy_baseline", "legacy_finetuned", "Swin-B", "ConvNeXtV2"]
MODEL_LABEL = {"legacy_baseline": "Legacy UNet baseline", "legacy_finetuned": "Legacy UNet finetuned",
               "Swin-B": "Swin-B (3-seed mean)", "ConvNeXtV2": "ConvNeXtV2 (3-seed mean)"}
METRICS = ["iou", "f1", "precision", "recall", "oa", "mcc"]
METRIC_LABEL = {"iou": "IoU", "f1": "F1", "precision": "Precision", "recall": "Recall",
                "oa": "Overall Accuracy", "mcc": "MCC"}
DATES = [20210305, 20210422, 20210621]

WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FONT = Font(bold=True, color="9C6500", size=12)
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
HDR_FONT = Font(bold=True)
METRIC_FILL = PatternFill("solid", fgColor="E2EFDA")
METRIC_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main() -> None:
    df = pd.read_csv(CSV)
    wb = openpyxl.load_workbook(REGISTRY)
    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(TAB)

    ncols = 2 + len(MODELS)  # date, wf, then one col per model

    # ---- banner ----
    banner = [
        "FOR CO-AUTHOR REPORTING ONLY — NOT FOR PUBLICATION",
        "Hampyeong Bay accuracy — reproduced with the DEPRECATED manuscript estimator",
        "Method: 400-pixel subsample, 10,000 bootstraps, 2.5/97.5 percentile CI. p-values OMITTED (the original's p-value was broken).",
        "This pixel-subsample method gives CIs too wide to separate models; it is superseded by the 2 km spatial block bootstrap.",
        "Defensible inference: experiments/hampyeong/evaluation/STATISTICAL_ASSESSMENT.md   |   Source: manuscript_estimator_all_metrics.csv",
    ]
    r = 1
    for i, line in enumerate(banner):
        c = ws.cell(row=r, column=1, value=line)
        c.font = WARN_FONT if i == 0 else Font(bold=(i == 1), italic=(i >= 2), color="9C6500")
        c.fill = WARN_FILL
        c.alignment = Alignment(wrap_text=False, vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
    r += 1

    # ---- one block per metric ----
    for metric in METRICS:
        mc = ws.cell(row=r, column=1, value=f"{METRIC_LABEL[metric]} — mean [2.5%, 97.5%] over 10,000×400-px bootstraps")
        mc.font = METRIC_FONT
        mc.fill = METRIC_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1

        # header row
        headers = ["Date", "GT water frac"] + [MODEL_LABEL[m] for m in MODELS]
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=j, value=h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
        r += 1

        for dt in DATES:
            wf = float(df[df.date == dt].water_frac.iloc[0])
            ws.cell(row=r, column=1, value=str(dt)).border = BORDER
            wfc = ws.cell(row=r, column=2, value=round(wf, 3)); wfc.border = BORDER; wfc.alignment = Alignment(horizontal="center")
            for j, m in enumerate(MODELS, start=3):
                row = df[(df.date == dt) & (df.model == m) & (df.metric == metric)].iloc[0]
                val = f"{row.boot_mean:.4f}  [{row.ci_lo:.4f}, {row.ci_hi:.4f}]"
                c = ws.cell(row=r, column=j, value=val)
                c.border = BORDER
                c.alignment = Alignment(horizontal="center")
            r += 1
        r += 1  # blank line between metric blocks

    # ---- footer ----
    footer = [
        "Notes:",
        "• New models (Swin-B, ConvNeXtV2) = 3-seed mean probability (seeds 19/42/58), stride 32, threshold 0.5.",
        "• Legacy = ResNet50-UNet from the deprecated sen12coast study (single run each): baseline (pretrained) and finetuned (site-tuned).",
        "• 3 dates are the only ones with a legacy prediction; the full 7-architecture / 5-date comparison is in STATISTICAL_ASSESSMENT.md.",
        "• MCC at 20210305 (94% water) is small-sample biased under 400-px subsampling — CIs there are unreliable by construction.",
        "• CIs here overlap across models on every date: this estimator cannot resolve the differences the block bootstrap confirms.",
    ]
    for i, line in enumerate(footer):
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(bold=(i == 0), italic=(i > 0), size=10, color="595959")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1

    # column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 26

    wb.save(REGISTRY)
    print(f"Added tab '{TAB}' to {REGISTRY}")
    print(f"Sheets now: {wb.sheetnames}")


if __name__ == "__main__":
    main()
